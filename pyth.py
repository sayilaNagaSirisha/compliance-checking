# app.py
import streamlit as st
import pandas as pd
import pdfplumber
import openpyxl
import re
import os

# To parse .docx files, you need to install python-docx
try:
    import docx
except ImportError:
    st.error("The 'python-docx' library is not installed. Please install it by running: pip install python-docx")
    st.stop()

# === Branding & Page Config ===
st.set_page_config(page_title="Regulatory Compliance & Safety Tool", layout="wide")

# --- FINAL, CORRECTED LOGO AND TITLE LAYOUT ---
col1, col2 = st.columns([1, 4])

with col1:
    def find_logo_path(possible_names=["logo.jpg", "logo.png", "logo.png.jpg"]):
        for name in possible_names:
            if os.path.exists(name):
                return name
        return None

    logo_path = find_logo_path()
    if logo_path:
        try:
            st.image(logo_path, width=150)
        except Exception as e:
            st.error(f"Error loading logo: {e}")
    else:
        st.warning("Logo not found.")

with col2:
    st.title("Regulatory Compliance & Safety Verification Tool")


# === Advanced CSS for Styling ===
st.markdown("""
<style>
.card{background:#f9f9f9; border-radius:10px; padding:15px; margin-bottom:10px; border-left: 5px solid #0056b3;}
.small-muted{color:#777; font-size:0.95em;}
.result-pass{color:#1e9f50; font-weight:700;}
.result-fail{color:#c43a31; font-weight:700;}
.main .block-container { padding-top: 2rem; }
</style>
""", unsafe_allow_html=True)

# === Session State Initialization ===
def init_session_state():
    state_defaults = {
        "reports_verified": 0, "requirements_generated": 0, "found_component": None,
        "component_db": pd.DataFrame(columns=['Part Number', 'Product Category', 'Manufacturer', 'Qualification', 'Voltage Rating DC', 'Dielectric', 'Capacitance', 'Tolerance'])
    }
    for key, value in state_defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value
init_session_state()

# === COMPLETE AND CORRECTED KNOWLEDGE BASE ===
TEST_CASE_KNOWLEDGE_BASE = {
    "water ingress": {
        "name": "Water Ingress Protection Test (IPX7)",
        "standard": "Based on ISO 20653 / IEC 60529",
        "description": "This test simulates the temporary immersion of the device in water to ensure no harmful quantity of water can enter the enclosure.",
        "procedure": [
            "Ensure the Device Under Test (DUT) is in a non-operational state and at ambient temperature.",
            "Submerge the DUT completely in a water tank.",
            "The lowest point of the DUT should be 1 meter below the surface of the water.",
            "The highest point of the DUT should be at least 0.15 meters below the surface.",
            "Maintain the immersion for the specified duration.",
            "After the test, remove the DUT, dry the exterior, and inspect the interior for any signs of water ingress.",
            "Conduct a functional check to ensure the device operates as expected."
        ],
        "parameters": {
            "Immersion Depth": "1 meter",
            "Test Duration": "30 minutes",
            "Water Temperature": "Ambient (within 5°C of DUT temperature)"
        },
        "equipment": ["Water Immersion Tank", "Depth Measurement Tool", "Stopwatch"]
    },
    "thermal shock": {
        "name": "Thermal Shock Test",
        "standard": "Based on ISO 16750-4",
        "description": "This test simulates the stress placed on electronic components when moving between extreme temperatures, such as a car being washed in winter.",
        "procedure": [
            "Set up a dual-chamber thermal shock system (hot and cold chambers).",
            "Place the DUT in the cold chamber and allow it to stabilize at the minimum temperature.",
            "Rapidly transfer the DUT to the hot chamber (transfer time should be less than 1 minute).",
            "Allow the DUT to stabilize at the maximum temperature.",
            "This completes one cycle. Repeat for the specified number of cycles.",
            "After the final cycle, allow the DUT to return to room temperature and perform a full functional and visual inspection."
        ],
        "parameters": {
            "Minimum Temperature": "-40°C",
            "Maximum Temperature": "+85°C",
            "Soak Time per Chamber": "1 hour",
            "Number of Cycles": "100 cycles"
        },
        "equipment": ["Dual-Chamber Thermal Shock System"]
    },
    "vibration": {
        "name": "Sinusoidal Vibration Test",
        "standard": "Based on IEC 60068-2-6",
        "description": "This test simulates the vibrations that a component might experience during its operational life due to engine harmonics or rough road conditions.",
        "procedure": [
            "Securely mount the DUT onto the vibration shaker table in its intended orientation.",
            "Sweep the frequency range from the minimum to the maximum value and back down.",
            "Perform the sweep on all three axes (X, Y, and Z).",
            "Maintain the specified G-force (acceleration) throughout the test.",
            "During the test, monitor the DUT for any intermittent failures or resonant frequencies.",
            "After the test, perform a full functional and visual inspection for any damage."
        ],
        "parameters": {
            "Frequency Range": "10 Hz to 500 Hz",
            "Acceleration": "5g (49 m/s²)",
            "Sweep Rate": "1 octave/minute",
            "Duration per Axis": "2 hours"
        },
        "equipment": ["Electrodynamic Shaker Table", "Vibration Controller", "Accelerometers"]
    },
    "short circuit": {
        "name": "External Short Circuit Protection",
        "standard": "Based on AIS-156 / IEC 62133-2",
        "description": "Verifies the safety performance of the battery or system when an external short circuit is applied.",
        "procedure": [
            "Ensure the DUT is fully charged.",
            "Connect the positive and negative terminals of the DUT with a copper wire or load with a resistance of less than 100 mΩ.",
            "Maintain the short circuit condition for the specified duration or until the protection circuit interrupts the current.",
            "Monitor the DUT for any hazardous events like fire, explosion, or casing rupture.",
            "Measure the case temperature during the test; it should not exceed the specified limit.",
            "After the test, the DUT should not show signs of fire or explosion."
        ],
        "parameters": {
            "Short Circuit Resistance": "< 100 mΩ",
            "Test Temperature": "55°C ± 5°C",
            "Observation Period": "1 hour after the event"
        },
        "equipment": ["High-Current Contactor", "Low-Resistance Load", "Thermocouples", "Safety Enclosure"]
    },
    "high temperature endurance": {
        "name": "High Temperature Endurance Test",
        "standard": "Based on ISO 16750-4 / IEC 60068-2-2",
        "description": "This test evaluates the device's ability to operate and survive prolonged exposure to high temperatures, simulating engine compartment conditions.",
        "procedure": [
            "Place the DUT in a high-temperature chamber.",
            "Bring the chamber temperature up to the specified maximum operating temperature.",
            "Maintain the temperature for the specified duration.",
            "During the test, apply nominal voltage and monitor the DUT for functional anomalies.",
            "After the test, allow the DUT to return to room temperature and perform a full functional and visual inspection."
        ],
        "parameters": {
            "Temperature": "+105°C (non-operating) / +85°C (operating)",
            "Duration": "96 hours to 1000 hours (depending on severity level)",
            "Relative Humidity": "Less than 20% RH (unless combined with humidity)"
        },
        "equipment": ["Temperature Chamber", "Data Logger", "Power Supply"]
    },
    "low temperature endurance": {
        "name": "Low Temperature Endurance Test",
        "standard": "Based on ISO 16750-4 / IEC 60068-2-1",
        "description": "This test assesses the device's performance and structural integrity under prolonged exposure to low temperatures, such as those found in cold climates.",
        "procedure": [
            "Place the DUT in a low-temperature chamber.",
            "Bring the chamber temperature down to the specified minimum operating temperature.",
            "Maintain the temperature for the specified duration.",
            "During the test, apply nominal voltage and monitor the DUT for functional anomalies.",
            "After the test, allow the DUT to return to room temperature and perform a full functional and visual inspection."
        ],
        "parameters": {
            "Temperature": "-40°C (non-operating) / -30°C (operating)",
            "Duration": "96 hours to 1000 hours (depending on severity level)"
        },
        "equipment": ["Low-Temperature Chamber", "Data Logger", "Power Supply"]
    },
    "temperature cycling": {
        "name": "Temperature Cycling Test",
        "standard": "Based on ISO 16750-4 / IEC 60068-2-14",
        "description": "This test subjects the device to alternating high and low temperatures to detect failures caused by thermal expansion and contraction, which can lead to material fatigue and joint failures.",
        "procedure": [
            "Place the DUT in a temperature cycling chamber.",
            "Cycle the temperature between the specified minimum and maximum values.",
            "Maintain the specified dwell time at each extreme temperature.",
            "Ensure the transition rate between temperatures is within specified limits.",
            "Perform the specified number of cycles.",
            "After the final cycle, allow the DUT to return to room temperature and perform a full functional and visual inspection."
        ],
        "parameters": {
            "Minimum Temperature": "-40°C",
            "Maximum Temperature": "+105°C",
            "Dwell Time at Extremes": "1 hour",
            "Transition Rate": "5°C/minute minimum",
            "Number of Cycles": "50 to 1000 cycles"
        },
        "equipment": ["Thermal Cycling Chamber", "Data Logger", "Power Supply (for operational checks)"]
    },
    "humidity & damp heat test": {
        "name": "Humidity & Damp Heat Test",
        "standard": "Based on ISO 16750-4 / IEC 60068-2-78 (Steady State) / IEC 60068-2-30 (Cyclic)",
        "description": "This test evaluates the device's resistance to high humidity, which can cause corrosion, insulation degradation, and moisture absorption in materials.",
        "procedure": [
            "Place the DUT in a climatic chamber.",
            "Set the chamber to the specified temperature and relative humidity.",
            "Maintain the conditions for the specified duration (steady state) or cycle through specified temperature and humidity profiles (cyclic).",
            "If specified, apply nominal power to the DUT during certain phases.",
            "Monitor for condensation, water droplets, and functional anomalies.",
            "After the test, allow the DUT to dry, then perform a full functional and visual inspection."
        ],
        "parameters": {
            "Temperature": "+40°C (Steady) / +25°C to +55°C (Cyclic)",
            "Relative Humidity": "93% RH (Steady) / 95% RH (Cyclic)",
            "Duration": "48 hours to 56 days (Steady) / 24 hours per cycle, 2-6 cycles (Cyclic)"
        },
        "equipment": ["Climatic Chamber (Humidity Chamber)", "Data Logger", "Humidity Sensor"]
    },
    "salt spray / corrosion test": {
        "name": "Salt Spray / Corrosion Test",
        "standard": "Based on ISO 9227 / ASTM B117",
        "description": "This test assesses the resistance of components and coatings to corrosion in a saline environment, simulating exposure to road salt or marine conditions.",
        "procedure": [
            "Prepare a salt solution (e.g., 5% NaCl solution).",
            "Place the DUT in a salt spray chamber.",
            "Atomize the salt solution into a fine mist within the chamber.",
            "Maintain the specified temperature and continuous salt spray for the duration.",
            "Periodically inspect the DUT for signs of corrosion (e.g., red rust, white rust).",
            "After the test, thoroughly rinse and dry the DUT, then perform a functional and visual inspection."
        ],
        "parameters": {
            "Salt Solution Concentration": "5% NaCl",
            "Temperature": "35°C",
            "pH of Solution": "6.5 to 7.2",
            "Test Duration": "24 hours to 1000 hours (depending on material/coating)"
        },
        "equipment": ["Salt Spray Chamber", "Salt Solution Preparation", "pH Meter"]
    },
    "dust ingress (ip rating)": {
        "name": "Dust Ingress Protection Test (IP Rating)",
        "standard": "Based on IEC 60529 (IP5X, IP6X)",
        "description": "This test verifies the protection of an enclosure against the ingress of dust, which can affect electrical components, moving parts, and optical surfaces.",
        "procedure": [
            "Place the DUT in a dust chamber filled with specified talcum powder (or similar fine dust).",
            "A vacuum pump may be connected to the DUT if negative pressure is required (for category 1 enclosures).",
            "Circulate the dust within the chamber using fans or blowers for the specified duration.",
            "After the test, remove the DUT and carefully clean the exterior.",
            "Disassemble the DUT (if applicable) and visually inspect the interior for any dust penetration.",
            "Conduct a functional check."
        ],
        "parameters": {
            "Dust Type": "Talcum powder (particle size < 75µm)",
            "Dust Concentration": "2 kg/m³",
            "Test Duration": "2 to 8 hours (IP5X) / 8 hours with vacuum (IP6X)",
            "Air Velocity": "Variable, to ensure uniform dust suspension"
        },
        "equipment": ["Dust Chamber", "Vacuum Pump (optional)", "Talcum Powder", "Air Circulator"]
    },
    "drop test / mechanical shock": {
        "name": "Drop Test / Mechanical Shock",
        "standard": "Based on IEC 60068-2-27 (Shock) / MIL-STD-810G Method 516.6 (Drop)",
        "description": "This test assesses the device's ability to withstand sudden, non-repetitive forces, such as those experienced during accidental drops, impacts, or rough handling.",
        "procedure": [
            "Securely mount the DUT onto a shock testing machine or prepare it for free fall.",
            "For shock: Subject the DUT to a specified number of shocks (half-sine or sawtooth pulse) along each of its three axes.",
            "For drop: Drop the DUT from a specified height onto a hard surface, ensuring it lands in different orientations.",
            "After each shock/drop, visually inspect the DUT for damage.",
            "Perform functional checks at specified intervals or after all shocks/drops.",
            "The DUT should remain functional and structurally sound."
        ],
        "parameters": {
            "Shock Pulse": "Half-sine",
            "Peak Acceleration": "50g to 200g (depending on application)",
            "Pulse Duration": "6 ms to 11 ms",
            "Number of Shocks": "3 to 18 shocks per axis",
            "Drop Height": "0.5 meters to 1.5 meters",
            "Number of Drops": "Multiple drops on faces, edges, corners"
        },
        "equipment": ["Shock Testing Machine", "Drop Tester", "Accelerometer", "Data Acquisition System"]
    },
    "overvoltage protection test": {
        "name": "Overvoltage Protection Test",
        "standard": "Based on ISO 16750-2 / LV 124",
        "description": "This test verifies that the device can withstand transient or continuous overvoltage conditions without damage or permanent degradation, simulating faults in the vehicle's electrical system.",
        "procedure": [
            "Connect the DUT to a programmable power supply.",
            "Apply the specified overvoltage level to the DUT's power input.",
            "Maintain the overvoltage for the specified duration.",
            "Monitor the DUT for smoke, fire, or catastrophic failure.",
            "After the test, return to nominal voltage and perform a functional check.",
            "The device's protection circuit should activate, or it should withstand the overvoltage without damage."
        ],
        "parameters": {
            "Test Voltage": "18V (for 12V systems) or higher transients",
            "Duration": "60 minutes (continuous) / < 1 second (transient)",
            "Test Temperature": "Room ambient"
        },
        "equipment": ["Programmable DC Power Supply", "Voltmeter", "Ammeter", "Load Box"]
    },
    "overcurrent protection test": {
        "name": "Overcurrent Protection Test",
        "standard": "Based on UL 60950-1 / IEC 62368-1",
        "description": "This test confirms that the device's protection mechanisms (e.g., fuses, current limiters) effectively prevent damage from excessive current draw, simulating a short circuit or overload.",
        "procedure": [
            "Connect the DUT to a power supply with current limiting capabilities.",
            "Gradually increase the current beyond the DUT's rated operating current.",
            "Observe at what current level the protection mechanism (fuse blows, circuit breaker trips, current limiter engages) activates.",
            "Verify that the device safely enters a protected state without generating excessive heat, smoke, or fire.",
            "After the protection activates, confirm that the device is not permanently damaged (unless it's a non-resettable fuse).",
            "Perform a functional check after the test (if applicable)."
        ],
        "parameters": {
            "Overcurrent Level": "1.5 times nominal current up to short circuit",
            "Activation Time": "Within specified limits (e.g., < 1 second for severe overload)",
            "Monitoring": "Current, Voltage, Temperature"
        },
        "equipment": ["Programmable Power Supply", "Electronic Load", "High-Speed Ammeter", "Thermal Camera"]
    },
    "insulation resistance test": {
        "name": "Insulation Resistance Test",
        "standard": "Based on IEC 60364-6 / ISO 6469-1",
        "description": "This test measures the electrical resistance of insulation materials to ensure they adequately prevent current leakage, which is crucial for safety and proper function, especially in high-voltage applications.",
        "procedure": [
            "Isolate the DUT from all power sources and ensure all capacitors are discharged.",
            "Connect the insulation resistance tester (megohmmeter) between the conductors and ground, or between different insulated conductors.",
            "Apply a specified DC test voltage (e.g., 500V or 1000V) for a set duration.",
            "Read and record the insulation resistance value.",
            "Repeat for all relevant insulation points.",
            "The measured resistance must exceed the specified minimum value."
        ],
        "parameters": {
            "Test Voltage": "500V DC or 1000V DC",
            "Minimum Resistance": "Generally > 1 MΩ (operating) / > 5 MΩ (new product)",
            "Test Duration": "1 minute"
        },
        "equipment": ["Insulation Resistance Tester (Megohmmeter)"]
    },
    "dielectric strength test": {
        "name": "Dielectric Strength (Hipot) Test",
        "standard": "Based on IEC 60950-1 / IEC 62368-1",
        "description": "This test applies a high voltage across insulation barriers to confirm they can withstand electrical stress without breaking down, preventing electric shock hazards.",
        "procedure": [
            "Isolate the DUT from all power sources.",
            "Connect a Hipot tester across the insulation barrier (e.g., between primary and secondary circuits, or between live parts and accessible conductive parts).",
            "Gradually increase the test voltage to the specified level (AC or DC).",
            "Maintain the test voltage for the specified duration.",
            "Monitor for breakdown (arc-over) or excessive leakage current.",
            "The insulation must withstand the voltage without breakdown or exceeding the current limit."
        ],
        "parameters": {
            "Test Voltage": "e.g., 1500V AC or 2121V DC (for basic insulation)",
            "Test Duration": "60 seconds or 1 second (production line)",
            "Leakage Current Limit": "Typically < 5 mA"
        },
        "equipment": ["Hipot Tester (Dielectric Withstand Tester)"]
    },
    "electrostatic discharge (esd) test": {
        "name": "Electrostatic Discharge (ESD) Test",
        "standard": "Based on ISO 10605 / IEC 61000-4-2",
        "description": "This test evaluates the device's immunity to electrostatic discharges, which can occur from human contact or charged objects, potentially causing malfunctions or damage to sensitive electronics.",
        "procedure": [
            "Place the DUT on a ground reference plane.",
            "Apply ESD pulses to various points on the DUT (contact discharge) and to nearby surfaces (air discharge).",
            "Perform both positive and negative polarity discharges.",
            "Monitor the DUT for any functional degradation, resets, or permanent damage during and after discharges.",
            "The device should either operate without interruption or recover to its normal state after the discharge."
        ],
        "parameters": {
            "Contact Discharge Voltage": "±2 kV to ±8 kV",
            "Air Discharge Voltage": "±2 kV to ±15 kV (depending on severity level)",
            "Number of Discharges": "10 per test point",
            "Repetition Rate": "Minimum 1 second between discharges"
        },
        "equipment": ["ESD Simulator Gun", "Ground Reference Plane", "Coupling Plane"]
    },
    "emi/emc test (electromagnetic compatibility)": {
        "name": "EMI/EMC Test (Electromagnetic Compatibility)",
        "standard": "Based on CISPR 25 (Emissions) / ISO 11452 (Immunity) / ECE R10",
        "description": "This is a broad category of tests ensuring the device does not interfere with other electronic systems (emissions) and is immune to external electromagnetic interference (immunity).",
        "procedure": [
            "This involves multiple sub-tests (Radiated Emissions, Conducted Emissions, Radiated Immunity, Conducted Immunity).",
            "Each sub-test has specific setup, instrumentation, and pass/fail criteria.",
            "Generally, the DUT is placed in an anechoic chamber or shielded room.",
            "Measurements are taken across a specified frequency range.",
            "The DUT must meet defined limits for emitted electromagnetic energy and operate without degradation when subjected to specified levels of interference."
        ],
        "parameters": {
            "Frequency Ranges": "Vary by sub-test (e.g., 150 kHz - 1 GHz)",
            "Limit Lines": "Defined by standard (e.g., dBµV/m for emissions, V/m for immunity)",
            "Modulation": "Specific modulations for immunity (e.g., AM, Pulse)"
        },
        "equipment": ["Anechoic Chamber", "EMI Receiver/Spectrum Analyzer", "Antennas", "Signal Generators", "Power Amplifiers"]
    },
    "conducted immunity test": {
        "name": "Conducted Immunity Test (CI)",
        "standard": "Based on ISO 11452-4 (BCI) / IEC 61000-4-6",
        "description": "This test subjects the device to radio-frequency disturbances injected directly into its cables, simulating interference from nearby cables or power lines.",
        "procedure": [
            "Place the DUT on a ground plane.",
            "Inject modulated RF signals onto the DUT's power and signal lines using a coupling clamp or CDNs (Coupling/Decoupling Networks).",
            "Sweep the frequency range at specified power levels.",
            "Monitor the DUT for any functional anomalies or degradation during the test.",
            "The device must maintain normal operation throughout the test without error or degradation below acceptable limits."
        ],
        "parameters": {
            "Frequency Range": "150 kHz to 200 MHz",
            "Test Level": "e.g., 100 mA (BCI) / 3 Vrms, 10 Vrms (IEC)",
            "Modulation": "80% AM, 1 kHz sine wave"
        },
        "equipment": ["RF Signal Generator", "Power Amplifier", "Coupling/Decoupling Networks (CDNs) or Bulk Current Injection (BCI) Probe", "Spectrum Analyzer"]
    },
    "radiated emissions test": {
        "name": "Radiated Emissions Test (RE)",
        "standard": "Based on CISPR 25 / ECE R10 / FCC Part 15",
        "description": "This test measures the electromagnetic fields radiated by the device into the air to ensure it does not cause interference with other electronic systems.",
        "procedure": [
            "Place the DUT in an anechoic chamber or on an open-area test site (OATS).",
            "Operate the DUT in its typical modes.",
            "Scan across a specified frequency range using calibrated antennas at defined measurement distances.",
            "Measure both horizontal and vertical polarization of the electric field.",
            "The measured emissions must fall below the specified limit lines.",
            "Identify and investigate any emissions exceeding limits."
        ],
        "parameters": {
            "Frequency Range": "30 MHz to 1 GHz (or higher)",
            "Measurement Distance": "1 meter, 3 meters, or 10 meters",
            "Limit Lines": "Defined in dBµV/m by the standard",
            "Antenna Types": "Biconical, Log-Periodic, Horn"
        },
        "equipment": ["Anechoic Chamber / OATS", "EMI Receiver / Spectrum Analyzer", "Antennas", "Preamplifiers", "Turntable", "Antenna Mast"]
    },
    "endurance / life cycle test": {
        "name": "Endurance / Life Cycle Test",
        "standard": "Generic, application-specific",
        "description": "This test simulates the total expected operational life of a component to identify potential long-term wear-out mechanisms or degradation over time.",
        "procedure": [
            "Operate the DUT continuously or through specified cycles (e.g., power on/off cycles, switch actuations, motor runs).",
            "Conduct the test under nominal environmental conditions (or accelerated conditions if specified).",
            "Periodically perform functional checks and measure key performance parameters.",
            "Record any failures, degradation, or changes in performance over the test duration.",
            "The device should maintain its specified performance throughout its expected lifespan."
        ],
        "parameters": {
            "Test Duration": "Equivalent to 5 to 15 years of operational life (e.g., 5000 hours, 100,000 cycles)",
            "Operating Conditions": "Nominal voltage, current, temperature",
            "Monitoring Frequency": "Regular intervals or continuous logging"
        },
        "equipment": ["Test Bench with DUT fixtures", "Programmable Controller (for cycling)", "Data Logger", "Measurement Instruments"]
    },
    "connector durability test": {
        "name": "Connector Durability Test",
        "standard": "Based on USCAR-2 / IEC 60512-9",
        "description": "This test assesses the mechanical and electrical integrity of electrical connectors over repeated mating and unmating cycles, simulating their operational lifespan.",
        "procedure": [
            "Mount the connector to a test fixture that simulates its application.",
            "Perform repeated mating and unmating cycles using an automated machine.",
            "Conduct the test at a specified speed and force.",
            "Periodically perform electrical measurements (e.g., contact resistance, insulation resistance) during or after a set number of cycles.",
            "Visually inspect the connector for mechanical damage (e.g., deformation, wear, breakage of contacts or housing).",
            "The contact resistance should remain stable and within limits, and no mechanical damage should occur."
        ],
        "parameters": {
            "Number of Cycles": "10 to 1000 cycles (depending on application)",
            "Mating/Unmating Force": "Measured or controlled by machine",
            "Contact Resistance Limit": "Typically < 10 mΩ",
            "Test Environment": "Room ambient (or combined with environmental tests)"
        },
        "equipment": ["Connector Mating/Unmating Machine", "Contact Resistance Meter", "Optical Inspection Tools"]
    }
}


# --- COMPLETE, UNIFIED, and FULLY POPULATED Component Database ---
UNIFIED_COMPONENT_DB = {
    "cga3e1x7r1e105k080ac": {"Manufacturer":"TDK", "Product Category":"Multilayer Ceramic Capacitors MLCC - SMD/SMT", "RoHS":"Yes", "Capacitance":"1 uF", "Voltage Rating DC":"25 VDC", "Dielectric":"X7R", "Tolerance":"10 %", "Case Code - in":"0603", "Case Code - mm":"1608", "Termination Style":"SMD/SMT", "Termination":"Standard", "Minimum Operating Temperature":"-55 C", "Maximum Operating Temperature":"+125 C", "Length":"1.6 mm", "Width":"0.8 mm", "Height":"0.8 mm", "Product":"Automotive MLCCs", "Qualification":"AEC-Q200"},
    "spc560p50l3": {"Manufacturer": "STMicroelectronics", "Product Category": "MCU", "RoHS": "Yes", "CPU Core": "PowerPC e200z0h", "Frequency": "64 MHz", "RAM Size": "48KB", "Flash Size": "512KB", "Package": "LQFP-100", "Minimum Operating Temperature": "-40 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q100"},
    "tja1051t": {"Manufacturer": "NXP", "Product Category": "CAN Transceiver", "RoHS": "Yes", "Data Rate": "1 Mbps", "Voltage Rating DC": "5V", "Package": "SO-8", "Minimum Operating Temperature": "-40 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q100"},
    "tle4275g": {"Manufacturer": "Infineon", "Product Category": "LDO Regulator", "RoHS": "Yes", "Output Voltage": "5V", "Output Current": "450mA", "Package": "TO-252-3", "Minimum Operating Temperature": "-40 C", "Maximum Operating Temperature": "150 C", "Qualification": "AEC-Q100"},
    "fsbb30ch60f": {"Manufacturer": "onsemi", "Product Category": "IGBT Module", "RoHS": "Yes", "Voltage Rating DC": "600V", "Current": "30A", "Package": "SPM27-CC", "Minimum Operating Temperature": "-40 C", "Maximum Operating Temperature": "150 C", "Product": "Smart Power Module"},
    "wslp2512r0100fe": {"Manufacturer": "Vishay", "Product Category": "Resistor", "RoHS": "Yes", "Resistance": "10 mOhm", "Power": "1W", "Tolerance": "1%", "Case Code - in": "2512", "Minimum Operating Temperature": "-65 C", "Maximum Operating Temperature": "170 C", "Qualification": "AEC-Q200"},
    "bq76952": {"Manufacturer": "Texas Instruments", "Product Category": "Battery Monitor", "RoHS": "Yes", "Cell Count": "3-16", "Interface": "I2C, SPI", "Package": "TQFP-48", "Minimum Operating Temperature": "-40 C", "Maximum Operating Temperature": "85 C", "Qualification": "AEC-Q100"},
    "irfz44n": {"Manufacturer": "Infineon", "Product Category": "MOSFET", "RoHS": "Yes", "Vds": "55V", "Id": "49A", "Rds(on)": "17.5 mOhm", "Package": "TO-220AB", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "175 C"},
    "1n4007": {"Manufacturer": "Multiple", "Product Category": "Diode", "RoHS": "Yes", "VRRM": "1000V", "If(AV)": "1A", "Package": "DO-41", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "150 C"},
    "fh28-10s-0.5sh(05)": {"Manufacturer": "Hirose", "Product Category": "Connector", "RoHS": "Yes", "Pitch": "0.5mm", "Positions": "10", "Current": "0.5A", "Package": "FFC/FPC", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "85 C"},
    "grt1555c1e220ja02j": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "22pF", "Voltage Rating DC": "25V", "Dielectric": "C0G", "Case Code - mm": "1005", "Tolerance": "5%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "grt155r61a475me13d": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "4.7uF", "Voltage Rating DC": "10V", "Dielectric": "X5R", "Case Code - mm": "1005", "Tolerance": "20%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "85 C", "Qualification": "AEC-Q200"},
    "grt31cr61a476ke13l": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "47uF", "Voltage Rating DC": "10V", "Dielectric": "X5R", "Case Code - mm": "3216", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "85 C", "Qualification": "AEC-Q200"},
    "cga2b2c0g1h180j050ba": {"Manufacturer": "TDK", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "18pF", "Voltage Rating DC": "50V", "Dielectric": "C0G", "Case Code - mm": "1005", "Tolerance": "5%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "c0402c103k4racauto": {"Manufacturer": "KEMET", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "10nF", "Voltage Rating DC": "16V", "Dielectric": "X7R", "Case Code - mm": "1005", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "gcm1555c1h101ja16d": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "100pF", "Voltage Rating DC": "50V", "Dielectric": "C0G", "Case Code - mm": "1005", "Tolerance": "5%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "grt155r71h104ke01d": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "0.1uF", "Voltage Rating DC": "50V", "Dielectric": "X7R", "Case Code - mm": "1005", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "grt21br61e226me13l": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "22uF", "Voltage Rating DC": "25V", "Dielectric": "X5R", "Case Code - mm": "2012", "Tolerance": "20%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "85 C", "Qualification": "AEC-Q200"},
    "grt1555c1h150fa02d": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "15pF", "Voltage Rating DC": "50V", "Dielectric": "C0G", "Case Code - mm": "1005", "Tolerance": "1%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "0402yc222j4t2a": {"Manufacturer": "AVX", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "2.2nF", "Voltage Rating DC": "16V", "Dielectric": "X7R", "Case Code - in": "0402", "Tolerance": "5%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "gcm1555c1h560fa16d": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "56pF", "Voltage Rating DC": "50V", "Dielectric": "C0G", "Case Code - mm": "1005", "Tolerance": "1%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "grt1555c1h330fa02d": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "33pF", "Voltage Rating DC": "50V", "Dielectric": "C0G", "Case Code - mm": "1005", "Tolerance": "1%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "grt188c81a106me13d": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "10uF", "Voltage Rating DC": "10V", "Dielectric": "X6S", "Case Code - mm": "1608", "Tolerance": "20%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "105 C", "Qualification": "AEC-Q200"},
    "umk212b7105kfna01": {"Manufacturer": "Taiyo Yuden", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "1uF", "Voltage Rating DC": "50V", "Dielectric": "X7R", "Case Code - in": "0805", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C"},
    "c1206c104k5racauto": {"Manufacturer": "KEMET", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "0.1uF", "Voltage Rating DC": "50V", "Dielectric": "X7R", "Case Code - in": "1206", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "grt31cr61h106ke01k": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "10uF", "Voltage Rating DC": "50V", "Dielectric": "X5R", "Case Code - in": "1206", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "85 C", "Qualification": "AEC-Q200"},
    "c0402c333k4racauto": {"Manufacturer": "KEMET", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "33nF", "Voltage Rating DC": "16V", "Dielectric": "X7R", "Case Code - in": "0402", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "cl10b474ko8vpnc": {"Manufacturer": "Samsung", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "0.47uF", "Voltage Rating DC": "16V", "Dielectric": "X7R", "Case Code - in": "0603", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C"},
    "gcm155r71c224ke02d": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "0.22uF", "Voltage Rating DC": "16V", "Dielectric": "X7R", "Case Code - mm": "1005", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "gcm155r71h102ka37j": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "1nF", "Voltage Rating DC": "50V", "Dielectric": "X7R", "Case Code - mm": "1005", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "50tpv330m10x10.5": {"Manufacturer": "Panasonic", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "330uF", "Voltage Rating DC": "50V", "Type": "Polymer", "ESR": "18 mOhm", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "105 C"},
    "cl31b684kbhwpne": {"Manufacturer": "Samsung", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "0.68uF", "Voltage Rating DC": "50V", "Dielectric": "X7R", "Case Code - in": "1206", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C"},
    "gcm155r71h272ka37d": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "2.7nF", "Voltage Rating DC": "50V", "Dielectric": "X7R", "Case Code - mm": "1005", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "edk476m050s9haa": {"Manufacturer": "KEMET", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "47uF", "Voltage Rating DC": "50V", "Type": "Aluminum Electrolytic", "ESR": "700 mOhm", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "105 C"},
    "gcm155r71h332ka37j": {"Manufacturer": "Murata", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "3.3nF", "Voltage Rating DC": "50V", "Dielectric": "X7R", "Case Code - mm": "1005", "Tolerance": "10%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "a768ke336m1hlae042": {"Manufacturer": "KEMET", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "33uF", "Voltage Rating DC": "50V", "Type": "Polymer", "ESR": "42 mOhm", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "ac0402jrx7r9bb152": {"Manufacturer": "Yageo", "Product Category": "Capacitor", "RoHS": "Yes", "Capacitance": "1.5nF", "Voltage Rating DC": "50V", "Dielectric": "X7R", "Case Code - in": "0402", "Tolerance": "5%", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q200"},
    "d5v0h1b2lpq-7b": {"Manufacturer": "Diodes Inc.", "Product Category": "TVS Diode", "RoHS": "Yes", "V Rwm": "5V", "Power": "30W", "Package": "X2-DFN1006-2", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "150 C"},
    "szmmbz9v1alt3g": {"Manufacturer": "onsemi", "Product Category": "Zener Diode", "RoHS": "Yes", "Vz": "9.1V", "Power": "225mW", "Tolerance": "5%", "Package": "SOT-23", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "150 C"},
    "d24v0s1u2tq-7": {"Manufacturer": "Diodes Inc.", "Product Category": "TVS Diode Array", "RoHS": "Yes", "V Rwm": "24V", "Channels": "1", "Package": "SOD-323", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "150 C"},
    "mmsz5225bt1g": {"Manufacturer": "onsemi", "Product Category": "Zener Diode", "RoHS": "Yes", "Vz": "3V", "Power": "500mW", "Tolerance": "5%", "Package": "SOD-123", "Minimum Operating Temperature": "-65 C", "Maximum Operating Temperature": "150 C"},
    "1n4148ws-7-f": {"Manufacturer": "Diodes Inc.", "Product Category": "Diode", "RoHS": "Yes", "Vrrm": "75V", "If(AV)": "150mA", "Package": "SOD-323", "Minimum Operating Temperature": "-65 C", "Maximum Operating Temperature": "150 C"},
    "ss34-e3/57t": {"Manufacturer": "Vishay", "Product Category": "Schottky Diode", "RoHS": "Yes", "Vrrm": "40V", "If(AV)": "3A", "Package": "DO-214AC", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "150 C"},
    "mbrm140t3g": {"Manufacturer": "onsemi", "Product Category": "Schottky Diode", "RoHS": "Yes", "Vrrm": "40V", "If(AV)": "1A", "Package": "SMA", "Minimum Operating Temperature": "-65 C", "Maximum Operating Temperature": "150 C"},
    "tps54331ddag": {"Manufacturer": "Texas Instruments", "Product Category": "DC/DC Converter", "RoHS": "Yes", "Topology": "Buck", "Input Voltage": "3.5V to 28V", "Output Voltage": "0.8V to 25V", "Output Current": "3A", "Package": "SOP-8", "Minimum Operating Temperature": "-40 C", "Maximum Operating Temperature": "150 C", "Qualification": "AEC-Q100"},
    "lm5066imm/nopb": {"Manufacturer": "Texas Instruments", "Product Category": "Hot Swap Controller", "RoHS": "Yes", "Input Voltage": "8V to 80V", "Package": "MSOP-10", "Minimum Operating Temperature": "-40 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q100"},
    "drr-34-86-7": {"Manufacturer": "Standex", "Product Category": "Reed Relay", "RoHS": "Yes", "Coil Voltage": "5V", "Contact Form": "SPST-NO", "Current Rating": "0.5A", "Package": "DIP", "Minimum Operating Temperature": "-40 C", "Maximum Operating Temperature": "125 C"},
    "si3445cdv-t1-ge3": {"Manufacturer": "Vishay", "Product Category": "MOSFET", "RoHS": "Yes", "Vds": "30V", "Id": "6.3A", "Rds(on)": "22 mOhm", "Package": "TSOP-6", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "150 C"},
    "dta114eekat146": {"Manufacturer": "Rohm Semiconductor", "Product Category": "Transistor", "RoHS": "Yes", "Type": "PNP Bipolar", "Ic": "-100mA", "Package": "SC-59", "Minimum Operating Temperature": "-55 C", "Maximum Operating Temperature": "150 C"},
    "lm2904qdrq1": {"Manufacturer": "Texas Instruments", "Product Category": "Operational Amplifier", "RoHS": "Yes", "Number of Channels": "2", "Voltage": "3V to 26V", "Package": "SOIC-8", "Minimum Operating Temperature": "-40 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q100"},
    "tps7a7805ddct": {"Manufacturer": "Texas Instruments", "Product Category": "LDO Regulator", "RoHS": "Yes", "Output Voltage": "5V", "Output Current": "50mA", "Input Voltage": "5V to 100V", "Package": "SOT-223", "Minimum Operating Temperature": "-40 C", "Maximum Operating Temperature": "125 C", "Qualification": "AEC-Q100"}
}

# --- UNIFIED HELPER FUNCTIONS (for data parsing and display) ---
def parse_uploaded_file(uploaded_file):
    """Parses various file types to extract text and data."""
    file_type = uploaded_file.type
    content = None
    if file_type == "application/pdf":
        with pdfplumber.open(uploaded_file) as pdf:
            content = " ".join(page.extract_text() for page in pdf.pages if page.extract_text())
    elif file_type == "text/plain":
        content = uploaded_file.getvalue().decode("utf-8")
    elif file_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        content = parse_xlsx(uploaded_file)
    elif file_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        content = parse_docx(uploaded_file)
    else:
        st.error(f"Unsupported file type: {file_type}")
        return None
    return content

def parse_docx(uploaded_file):
    """Parses a .docx file and returns its text content."""
    try:
        doc = docx.Document(uploaded_file)
        return "\n".join([p.text for p in doc.paragraphs])
    except Exception as e:
        st.error(f"An error occurred while parsing the DOCX file: {e}")
        return None

def parse_xlsx(uploaded_file):
    """Parses a .xlsx file and returns its text content."""
    try:
        workbook = openpyxl.load_workbook(uploaded_file)
        full_text = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            full_text.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows():
                row_text = [str(cell.value) if cell.value is not None else "" for cell in row]
                full_text.append("\t".join(row_text))
        return "\n".join(full_text)
    except Exception as e:
        st.error(f"An error occurred while parsing the XLSX file: {e}")
        return None

def extract_test_data(text_content):
    """
    Extracts individual test cases and their results from a continuous text block.
    This version uses a regex to find numbered test cases and their associated results.
    It can handle both the `XX: ... PASS` format and `XX: ... FAIL` format, ensuring
    each result is captured independently.
    """
    extracted_data = []

    # This regex is specifically designed to handle the format in your image:
    # `[number]: [Test Description] -> [Result]` or `[number]: [Test Description] "FAIL"`
    test_pattern = re.compile(
        r"(\d+): (.*?)(?:->| |)(PASS|FAIL|N/A|COMPLETE|SUCCESS|FAILURE)",
        re.IGNORECASE
    )

    matches = test_pattern.findall(text_content)

    if not matches:
        # Fallback to a simpler line-by-line search for results if the primary regex fails.
        # This can be useful for reports that aren't numbered.
        lines = text_content.split('\n')
        for line in lines:
            line = line.strip()
            if not line:
                continue

            result = None
            if "FAIL" in line.upper() or "FAILURE" in line.upper():
                result = "FAIL"
            elif "PASS" in line.upper() or "SUCCESS" in line.upper():
                result = "PASS"
            elif "N/A" in line.upper() or "NA" in line.upper():
                result = "N/A"

            if result:
                extracted_data.append({
                    "Test Description": line,
                    "Result": result
                })
    else:
        # Process the matches found by the regex
        for match in matches:
            test_id, description, result = match
            extracted_data.append({
                "Test Description": f"{test_id}: {description.strip()}",
                "Result": result.upper()
            })

    return extracted_data

def find_component_in_db(component_part_number):
    """
    Finds a component in the UNIFIED_COMPONENT_DB.
    Returns the component's dictionary if found, otherwise None.
    """
    normalized_part_number = component_part_number.lower().strip()
    return UNIFIED_COMPONENT_DB.get(normalized_part_number)

def display_test_card(test_data, color="#0056b3"):
    """Displays a single test case in a stylish card format."""
    st.markdown(f"""
    <div class="card" style="border-left: 5px solid {color};">
        <p><strong>Test:</strong> {test_data.get('Test Description', 'N/A')}</p>
        <p><strong>Result:</strong> <span style="color:{color}; font-weight: bold;">{test_data.get('Result', 'N/A')}</span></p>
    </div>
    """, unsafe_allow_html=True)


# --- MAIN APPLICATION LOGIC ---
st.sidebar.header("Navigation", anchor=False)
option = st.sidebar.radio("Select a Module", ("Regulatory Requirements", "Report Verification", "Component Information", "Dashboard & Analytics"))

# --- Regulatory Requirements Module ---
if option == "Regulatory Requirements":
    st.subheader("Regulatory Requirements Generator", anchor=False)
    st.caption("Generate detailed test requirements based on industry standards.")

    test_case_options = sorted(TEST_CASE_KNOWLEDGE_BASE.keys())
    selected_test_cases = st.multiselect(
        "Select the required tests:",
        options=test_case_options,
        help="Choose one or more test cases to generate detailed requirements."
    )

    if st.button("Generate Requirements"):
        st.session_state.requirements_generated += 1
        with st.spinner("Generating requirements..."):
            st.success("Requirements generated!")
            for test in selected_test_cases:
                test_info = TEST_CASE_KNOWLEDGE_BASE.get(test)
                if test_info:
                    st.markdown(f"### {test_info['name']}")
                    st.markdown(f"**Standard:** {test_info['standard']}")
                    st.markdown(f"**Description:** {test_info['description']}")
                    st.markdown("**Procedure:**")
                    for step in test_info['procedure']:
                        st.markdown(f"* {step}")
                    st.markdown("**Parameters:**")
                    for param, value in test_info['parameters'].items():
                        st.markdown(f"* **{param}:** {value}")
                    st.markdown("**Equipment:**")
                    st.markdown(f"* {', '.join(test_info['equipment'])}")
                    st.markdown("---")


# --- Report Verification Module ---
elif option == "Report Verification":
    st.subheader("Automated Report Verification", anchor=False)
    st.caption("Upload a test report to automatically identify PASS/FAIL results.")
    uploaded_file = st.file_uploader("Choose a file (PDF, TXT, DOCX, XLSX)", type=["pdf", "txt", "docx", "xlsx"])

    if uploaded_file:
        st.session_state.reports_verified += 1
        with st.spinner("Parsing and analyzing the report..."):
            text_content = parse_uploaded_file(uploaded_file)

        if text_content:
            parsed_data = extract_test_data(text_content)

            if parsed_data:
                passed = [t for t in parsed_data if "PASS" in str(t.get("Result", "")).upper()]
                failed = [t for t in parsed_data if "FAIL" in str(t.get("Result", "")).upper()]
                others = [t for t in parsed_data if not ("PASS" in str(t.get("Result", "")).upper() or "FAIL" in str(t.get("Result", "")).upper())]

                st.markdown(f"### Found {len(passed)} Passed, {len(failed)} Failed, and {len(others)} Other items.")

                if failed:
                    st.error(f"Report analysis complete. {len(failed)} FAILED test cases found.")
                    with st.expander("🔴 Failed Cases", expanded=True):
                        for t in failed: display_test_card(t, '#c43a31')
                else:
                    st.success("Report analysis complete. No FAILED test cases found.")

                if passed:
                    with st.expander("✅ Passed Cases", expanded=False):
                        for t in passed: display_test_card(t, '#1e9f50')

                if others:
                    with st.expander("ℹ️ Other/Informational Items", expanded=False):
                        for t in others: display_test_card(t, '#808080')
            else:
                st.warning("No recognizable test data was extracted. Please ensure the report contains clear PASS/FAIL keywords or numbered lists.")
        else:
            st.error("Failed to extract content from the uploaded file.")


# --- Component Information Module ---
elif option == "Component Information":
    st.subheader("Component Information Lookup", anchor=False)
    st.caption("Look up key specifications for automotive-grade components.")

    component_part_number = st.text_input("Enter Component Part Number:")

    if component_part_number:
        found_component_data = find_component_in_db(component_part_number)

        if found_component_data:
            # Create a DataFrame from the component data
            component_df = pd.DataFrame(found_component_data.items(), columns=['Property', 'Value'])
            component_df.set_index('Property', inplace=True)

            # Append to session state history if not already there
            normalized_part = component_part_number.lower()
            if normalized_part not in st.session_state.component_db['Part Number'].str.lower().tolist():
                new_row = pd.DataFrame([found_component_data])
                new_row.insert(0, 'Part Number', component_part_number)
                # Re-order columns to match the session state DataFrame
                new_row = new_row[st.session_state.component_db.columns]
                st.session_state.component_db = pd.concat([st.session_state.component_db, new_row], ignore_index=True)

            st.success(f"Component '{component_part_number}' found.")
            st.markdown("### Component Details:")
            st.dataframe(component_df, use_container_width=True)
        else:
            st.error(f"Component '{component_part_number}' not found in the database.")


# --- Dashboard & Analytics Module ---
elif option == "Dashboard & Analytics":
    st.subheader("Dashboard & Analytics", anchor=False)
    st.caption("High-level view of session activities.")
    c1, c2, c3 = st.columns(3)

    c1.metric("Reports Verified", st.session_state.reports_verified)
    c2.metric("Requirements Generated", st.session_state.requirements_generated)
    c3.metric("Components Looked Up", len(st.session_state.component_db))

    if not st.session_state.component_db.empty:
        st.markdown("### Recent Component Lookups")
        st.dataframe(st.session_state.component_db, use_container_container=True)